#!/usr/bin/env python3
import pandas as pd
import requests
import logging
import os
import sys
from datetime import datetime
import re
try:
    import openpyxl
except ImportError:
    logging.warning("Module openpyxl non trouvé. Installation recommandée pour le formatage conditionnel.")

# Configuration du logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler("enrich_siren.log"),
        logging.StreamHandler(sys.stdout)
    ]
)

def clean_address(address):
    """Nettoie et normalise une adresse pour la comparaison"""
    if not address or pd.isna(address):
        return ""
        
    address = str(address).upper()
    
    # Remplacer les abréviations courantes
    address = re.sub(r'\bBOULEVARD\b', 'BD', address)
    address = re.sub(r'\bAVENUE\b', 'AV', address)
    address = re.sub(r'\bROUTE\b', 'RTE', address)
    address = re.sub(r'\bRUE\b', 'R', address)
    address = re.sub(r'\bSAINT\b', 'ST', address)
    address = re.sub(r'\bSAINTE\b', 'STE', address)
    
    # Supprimer les caractères spéciaux et harmoniser les espaces
    address = re.sub(r'[^\w\s]', ' ', address)
    address = re.sub(r'\s+', ' ', address).strip()
    
    return address

def addresses_match(original_address, api_address):
    """Vérifie si deux adresses correspondent"""
    # Si l'une des adresses est vide ou non définie, pas de correspondance
    if not original_address or not api_address:
        return False
        
    # Si l'adresse originale est juste un point ou un caractère spécial isolé
    if original_address.strip() in ["·", ".", "-", "/"]:
        return False
    
    # Nettoyer les deux adresses
    norm_original = clean_address(original_address)
    norm_api = clean_address(api_address)
    
    # Si l'une des adresses normalisées est vide après nettoyage
    if not norm_original or not norm_api:
        return False
    
    # Vérifier si une adresse est contenue dans l'autre
    if norm_original in norm_api or norm_api in norm_original:
        return True
        
    # Comparer les mots significatifs
    original_words = set(norm_original.split())
    api_words = set(norm_api.split())
    common_words = original_words.intersection(api_words)
    
    # Extraire les codes postaux
    original_cp = re.search(r'\b\d{5}\b', original_address)
    api_cp = re.search(r'\b\d{5}\b', api_address)
    
    # Si les codes postaux sont différents, pas de correspondance
    if original_cp and api_cp and original_cp.group(0) != api_cp.group(0):
        return False
        
    # Si au moins 30% des mots correspondent et il y a au moins 2 mots en commun
    return len(common_words) >= max(2, len(original_words) * 0.3)

def search_company(company_name, address, token):
    """Recherche une entreprise dans l'API Recherche d'Entreprises"""
    logging.info(f"Recherche pour: {company_name}, adresse: {address}")
    
    # Traiter le nom de l'entreprise pour améliorer les correspondances
    simplified_name = company_name
    
    # Liste des villes françaises courantes qui pourraient apparaître dans les noms d'entreprises
    cities = ["PARIS", "LYON", "MARSEILLE", "TOULOUSE", "NICE", "NANTES", "MONTPELLIER", 
             "STRASBOURG", "BORDEAUX", "LILLE", "RENNES", "REIMS", "TOULON", "GRENOBLE", 
             "DIJON", "ANGERS", "NÎMES", "VILLEURBANNE", "SAINT-DENIS", "ASNIÈRES", "CAEN",
             "SAINT-ÉTIENNE", "ROUEN", "NANCY", "ORLÉANS", "LIMOGES", "MULHOUSE", 
             "SAINT-PAUL", "ROUBAIX", "DUNKERQUE", "PERPIGNAN", "AMIENS", "BOULOGNE", 
             "BESANÇON", "BREST", "CANNES", "METZ", "ANTIBES", "HONFLEUR", "HYMER", "FECAMP"]
    
    # Supprimer le texte après les séparateurs courants pour améliorer la recherche
    for separator in [' - ', ' | ', ' – ', ' – ', ' : ', ' / ']:
        if separator in simplified_name:
            simplified_name = simplified_name.split(separator)[0].strip()
            logging.debug(f"Nom simplifié (séparateur): {simplified_name}")
            break
    
    # Détecter et supprimer les noms de villes à la fin du nom
    words = simplified_name.upper().split()
    if len(words) > 1:
        # Vérifier si le ou les derniers mots correspondent à une ville
        if words[-1] in cities or ' '.join(words[-2:]).upper() in cities or (len(words) > 2 and ' '.join(words[-3:]).upper() in cities):
            # Si un des derniers mots est "ST" suivi d'un nom de ville
            if len(words) > 2 and words[-2] in ["ST", "SAINT"] and words[-1] in cities:
                # Garder "ST" et le nom de ville si c'est le nom de l'entreprise
                if len(words) > 3:
                    simplified_name = ' '.join(words[:-2]).strip()
                    logging.debug(f"Nom simplifié (ville avec ST): {simplified_name}")
            else:
                # Enlever juste le nom de la ville
                for i in range(1, min(4, len(words))):
                    city_candidate = ' '.join(words[-i:]).upper()
                    if city_candidate in cities:
                        simplified_name = ' '.join(words[:-i]).strip()
                        logging.debug(f"Nom simplifié (ville): {simplified_name}")
                        break
    
    headers = {
        "Authorization": f"Bearer {token}",
        "Accept": "application/json"
    }
    
    # Construire l'URL de recherche
    base_url = "https://recherche-entreprises.api.gouv.fr/search"
    params = {
        "q": simplified_name,
        "page": 1,
        "per_page": 5,
        "minimal": True,
        "include": "siege,dirigeants"
    }
    
    try:
        # Faire la requête à l'API
        response = requests.get(base_url, headers=headers, params=params, timeout=30)
        
        if response.status_code == 200:
            data = response.json()
            results = data.get('results', [])
            
            if not results:
                logging.warning(f"✗ Aucun résultat pour {company_name}")
                
                # Si l'adresse est disponible, faire une recherche secondaire par adresse
                if address and not address.strip() in ["·", ".", "-", "/"]:
                    logging.info(f"Tentative de recherche secondaire par adresse pour {company_name}")
                    # Extraire les éléments numériques de l'adresse pour la recherche
                    address_parts = address.split()
                    address_search = " ".join([p for p in address_parts if any(c.isdigit() for c in p) or p.lower() in ["rue", "avenue", "boulevard", "bd", "av", "rte", "route", "allée", "all", "chemin", "impasse", "place", "quai"]])
                    
                    if address_search.strip():
                        params["q"] = address_search
                        logging.debug(f"Recherche par adresse: {address_search}")
                        address_response = requests.get(base_url, headers=headers, params=params, timeout=30)
                        
                        if address_response.status_code == 200:
                            address_data = address_response.json()
                            address_results = address_data.get('results', [])
                            
                            if address_results:
                                logging.info(f"✓ {len(address_results)} résultat(s) trouvé(s) par recherche d'adresse")
                                
                                # Fonction pour calculer la similarité entre deux chaînes
                                def string_similarity(s1, s2):
                                    s1 = s1.upper()
                                    s2 = s2.upper()
                                    # Nombre de caractères en commun
                                    common = sum(c1 == c2 for c1, c2 in zip(s1, s2))
                                    # Plus la longueur est similaire, meilleur est le score
                                    length_diff = 1 - abs(len(s1) - len(s2)) / max(len(s1), len(s2))
                                    # Score combiné
                                    return common + length_diff * 5
                                
                                # Mots-clés liés à l'installation électrique
                                electrical_keywords = ["ELEC", "ELECTR", "ELECTRIC", "ELECTRONIQUE", "CÂBLAGE", "CABLAGE", 
                                                      "CABL", "INSTAL", "COURANT", "TELECOM", "ENERGI", "ENERGIE"]
                                
                                # Trouver le résultat le plus pertinent
                                best_match = None
                                best_score = -1
                                best_activity_match = None
                                best_activity_score = -1
                                
                                for result in address_results:
                                    result_name = result.get('nom_complet', '')
                                    description = result.get('objet_social', '') or result.get('description', '')
                                    activite = result.get('activite_principale', '') or ''
                                    score = string_similarity(company_name, result_name)
                                    
                                    # Vérifier si le nom ou la description contient des mots-clés liés à l'électricité
                                    has_electrical_keyword = False
                                    combined_text = (result_name + ' ' + description + ' ' + activite).upper()
                                    for keyword in electrical_keywords:
                                        if keyword in combined_text:
                                            has_electrical_keyword = True
                                            break
                                    
                                    # Si le score est suffisamment bon (nom similaire)
                                    if score > 10 and score > best_score:
                                        best_score = score
                                        best_match = result
                                    
                                    # Garder trace du meilleur match basé sur l'activité, même si le nom n'est pas similaire
                                    if has_electrical_keyword and (best_activity_match is None or score > best_activity_score):
                                        best_activity_score = score
                                        best_activity_match = result
                                
                                # Décider quel résultat utiliser
                                selected_match = None
                                match_reason = ""
                                
                                if best_match:
                                    # Si nous avons trouvé un nom suffisamment similaire
                                    selected_match = best_match
                                    match_reason = "similarité de nom"
                                elif best_activity_match:
                                    # Sinon, utiliser la correspondance basée sur l'activité
                                    selected_match = best_activity_match
                                    match_reason = "activité similaire"
                                
                                if selected_match:
                                    results = [selected_match]
                                    logging.info(f"✓ Meilleure correspondance par adresse ({match_reason}): {selected_match.get('nom_complet', '')}")
                                else:
                                    logging.warning(f"✗ Aucune correspondance pertinente trouvée via l'adresse pour {company_name}")
                                    return None
                            else:
                                return None
                        else:
                            logging.error(f"Erreur API recherche par adresse ({address_response.status_code}): {address_response.text}")
                            return None
                    else:
                        return None
                else:
                    return None
                
            # Chercher la meilleure correspondance
            best_match = None
            best_match_score = 0
            
            for result in results:
                # Vérifier la correspondance d'adresse
                api_address = result.get('siege', {}).get('adresse', '')
                match = addresses_match(address, api_address)
                
                # Si c'est une correspondance et qu'on n'a pas encore trouvé de correspondance
                if match and best_match is None:
                    best_match = result
                    best_match_score = 1
                    break
            
            # Si aucune correspondance trouvée, prendre le premier résultat
            if best_match is None and results:
                best_match = results[0]
                
            if best_match:
                # Construire les informations du dirigeant
                dirigeants_str = ""
                dirigeants = best_match.get('dirigeants', [])
                
                # Variables pour le dirigeant le plus jeune
                nom_dirigeant = ""
                prenom_dirigeant = ""
                qualite_dirigeant = ""
                annee_naissance_max = "0"
                
                # Liste pour stocker tous les dirigeants
                autres_dirigeants = []
                
                if dirigeants:
                    dirigeants_info = []
                    
                    # Trouver le dirigeant le plus jeune et collecter les informations sur tous les dirigeants
                    for dirigeant in dirigeants:
                        # Récupérer les différentes formes possibles des données
                        annee = dirigeant.get('annee_de_naissance', dirigeant.get('annee_naissance', ''))
                        nom = dirigeant.get('nom', '')
                        prenoms = dirigeant.get('prenoms', dirigeant.get('prenom', ''))
                        qualite = dirigeant.get('qualite', '')
                        type_dirigeant = dirigeant.get('type_dirigeant', '')
                        
                        # S'assurer que les chaînes sont des chaînes de caractères avant d'utiliser strip()
                        nom = nom.strip() if isinstance(nom, str) else ''
                        prenoms = prenoms.strip() if isinstance(prenoms, str) else ''
                        qualite = qualite.strip() if isinstance(qualite, str) else ''
                        
                        # Extraire uniquement le premier prénom pour éviter les doublons
                        premier_prenom = prenoms.split()[0] if prenoms and ' ' in prenoms else prenoms
                        
                        # Déterminer si c'est une personne morale
                        est_personne_morale = type_dirigeant == "personne_morale" or (prenoms == '' and ('SA' in nom or 'SAS' in nom or 'SARL' in nom or 'SCI' in nom or 'EURL' in nom))
                        
                        # Créer un dictionnaire avec les informations du dirigeant
                        dirigeant_info = {
                            'nom': nom,
                            'prenoms': prenoms,
                            'premier_prenom': premier_prenom,
                            'qualite': qualite,
                            'annee': annee,
                            'age': '',
                            'est_personne_morale': est_personne_morale,
                            'identifiant': f"{nom}_{premier_prenom}"  # Identifiant unique basé sur nom et premier prénom
                        }
                        
                        # Si le nom contient des parenthèses, extraire uniquement le nom principal
                        if nom and '(' in nom and ')' in nom:
                            nom = nom.split('(')[0].strip()
                            dirigeant_info['nom'] = nom
                            dirigeant_info['identifiant'] = f"{nom}_{premier_prenom}"
                        
                        # Calculer l'âge si année de naissance disponible
                        if annee:
                            try:
                                annee_actuelle = datetime.now().year
                                dirigeant_info['age'] = f"{annee_actuelle - int(annee)} ans"
                            except (ValueError, TypeError):
                                pass
                        
                        # Vérifier si ce dirigeant (même nom et premier prénom) existe déjà dans la liste
                        dirigeant_existant = False
                        for dir_existant in autres_dirigeants:
                            if dir_existant['identifiant'] == dirigeant_info['identifiant']:
                                dirigeant_existant = True
                                # Si le dirigeant existe déjà mais avec moins d'informations, mettre à jour
                                if len(dir_existant['prenoms']) < len(dirigeant_info['prenoms']):
                                    dir_existant.update(dirigeant_info)
                                break
                        
                        # Ajouter à la liste des dirigeants uniquement s'il n'existe pas déjà
                        if not dirigeant_existant:
                            autres_dirigeants.append(dirigeant_info)
                        
                        # Prendre le premier dirigeant si aucun n'est trouvé
                        if not nom_dirigeant and not prenom_dirigeant and nom:
                            nom_dirigeant = nom
                            prenom_dirigeant = premier_prenom
                            qualite_dirigeant = qualite
                        
                        # Si l'année est présente et plus récente (dirigeant plus jeune)
                        if annee and (not annee_naissance_max or annee > annee_naissance_max):
                            annee_naissance_max = annee
                            nom_dirigeant = nom
                            prenom_dirigeant = premier_prenom
                            qualite_dirigeant = qualite
                            
                        # Construire l'info complète pour la liste des dirigeants
                        type_str = "[PM] " if est_personne_morale else ""
                        nom_str = nom if nom else ''
                        prenom_str = prenoms if prenoms else ''
                        annee_str = f"({annee})" if annee else ''
                        qualite_str = f"- {qualite}" if qualite else ''
                        
                        info = f"{type_str}{nom_str} {prenom_str} {annee_str} {qualite_str}".strip()
                        # Éviter les espaces multiples
                        info = ' '.join(info.split())
                        if info:  # Ne pas ajouter de chaînes vides
                            dirigeants_info.append(info)
                    
                    dirigeants_str = " | ".join(dirigeants_info)
                
                # Vérifier la correspondance d'adresse
                api_address = best_match.get('siege', {}).get('adresse', '')
                address_match = addresses_match(address, api_address)
                
                # Récupérer la tranche d'effectif correctement
                tranche_effectif_code = best_match.get('siege', {}).get('tranche_effectif_salarie', '')
                
                # Convertir le code tranche d'effectif en description lisible
                tranche_effectif_map = {
                    "NN": "Unité non-employeuse ou présumée non-employeuse",
                    "00": "0 salarié (a employé des salariés)",
                    "01": "1 ou 2 salariés",
                    "02": "3 à 5 salariés",
                    "03": "6 à 9 salariés",
                    "11": "10 à 19 salariés",
                    "12": "20 à 49 salariés",
                    "21": "50 à 99 salariés",
                    "22": "100 à 199 salariés",
                    "31": "200 à 249 salariés",
                    "32": "250 à 499 salariés",
                    "41": "500 à 999 salariés",
                    "42": "1000 à 1999 salariés",
                    "51": "2000 à 4999 salariés",
                    "52": "5000 à 9999 salariés",
                    "53": "10000 salariés et plus"
                }
                
                tranche_effectif_desc = tranche_effectif_map.get(tranche_effectif_code, "null")
                
                # Calculer l'âge du dirigeant le plus jeune s'il y a une année de naissance
                age = ""
                if annee_naissance_max and annee_naissance_max != "0":
                    try:
                        annee_actuelle = datetime.now().year
                        age = f"{annee_actuelle - int(annee_naissance_max)} ans"
                    except (ValueError, TypeError):
                        age = ""
                
                # Préparer la liste des autres dirigeants à retourner
                autres_dirigeants_result = []
                
                # Filtrer le dirigeant principal de la liste des autres dirigeants
                autres_dirigeants_filtres = [d for d in autres_dirigeants if not (
                    d['nom'] == nom_dirigeant and 
                    d['premier_prenom'] == prenom_dirigeant
                )]
                
                # Limiter à 5 dirigeants alternatifs maximum
                for i in range(min(5, len(autres_dirigeants_filtres))):
                    dirigeant = autres_dirigeants_filtres[i]
                    autres_dirigeants_result.append({
                        'nom': dirigeant['nom'],
                        'prenoms': dirigeant['premier_prenom'],  # Utiliser uniquement le premier prénom
                        'qualite': dirigeant['qualite'],
                        'age': dirigeant['age'],
                        'est_personne_morale': dirigeant['est_personne_morale']
                    })
                
                # Extraire l'année de création (seulement l'année, pas le jour et le mois)
                annee_creation = ""
                date_creation_complete = best_match.get('date_creation', '')
                if date_creation_complete and len(date_creation_complete) >= 4:
                    annee_creation = date_creation_complete[:4]  # Prendre les 4 premiers caractères (l'année)
                
                return {
                    'siren': best_match.get('siren', ''),
                    'nom_raison_sociale': best_match.get('nom_complet', ''),
                    'adresse': api_address,
                    'etat_administratif': best_match.get('etat_administratif', ''),
                    'tranche_effectif_code': tranche_effectif_code,
                    'tranche_effectif': tranche_effectif_desc,
                    'date_creation': annee_creation,  # Afficher seulement l'année
                    'annee_creation': annee_creation,
                    'nom_dirigeant': nom_dirigeant,
                    'prenom_dirigeant': prenom_dirigeant,
                    'qualite_dirigeant': qualite_dirigeant,
                    'age_dirigeant': age,
                    'match_adresse': "Oui" if address_match else "Non",
                    'autres_dirigeants': autres_dirigeants_result
                }
            
            logging.warning(f"✗ Aucune correspondance pertinente pour {company_name}")
            return None
            
        else:
            logging.error(f"Erreur API ({response.status_code}): {response.text}")
            return None
            
    except Exception as e:
        logging.error(f"Exception lors de la recherche: {str(e)}")
        return None

def enrich_excel_file(input_file, output_file, token):
    """Enrichit un fichier Excel avec des données d'entreprises"""
    try:
        # Vérifier l'existence du fichier
        if not os.path.exists(input_file):
            logging.error(f"Le fichier {input_file} n'existe pas")
            return False
            
        # Lire le fichier Excel
        try:
            df = pd.read_excel(input_file)
            logging.info(f"Fichier Excel chargé: {len(df)} lignes")
            logging.info(f"Colonnes: {', '.join(df.columns)}")
            
            # Supprimer les colonnes "Nom", "Unnamed: 8" et "Siren" si elles existent
            if 'Nom' in df.columns:
                df = df.drop(columns=['Nom'])
                logging.info("Colonne 'Nom' supprimée car inutile")
                
            if 'Unnamed: 8' in df.columns:
                df = df.drop(columns=['Unnamed: 8'])
                logging.info("Colonne 'Unnamed: 8' supprimée car inutile")
                
            # Supprimer la colonne 'Siren' avec S majuscule si elle existe car elle est redondante avec 'SIREN'
            if 'Siren' in df.columns:
                df = df.drop(columns=['Siren'])
                logging.info("Colonne 'Siren' supprimée car redondante avec 'SIREN'")
        except Exception as e:
            logging.error(f"Erreur lors de la lecture du fichier: {str(e)}")
            return False
            
        # Vérifier les colonnes requises
        required_columns = ['Company', 'Address']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            logging.error(f"Colonnes manquantes: {', '.join(missing_columns)}")
            return False
            
        # Préparer les colonnes de résultats (sans CA ni Capital que l'API ne fournit pas)
        result_columns = ['SIREN', 'Nom_Raison_Sociale', 'Adresse', 'Etat_Administratif', 
                        'Tranche_Effectif', 'Date_Creation', 'Annee_Creation',
                        'Nom_Dirigeant', 'Prenom_Dirigeant', 'Qualite_Dirigeant', 'Age_Dirigeant', 'Match_Adresse']
        
        # Ajouter des colonnes pour les dirigeants alternatifs (5 maximum)
        for i in range(1, 6):
            result_columns.extend([
                f'Dirigeant{i}_Nom',
                f'Dirigeant{i}_Prenom',
                f'Dirigeant{i}_Qualite',
                f'Dirigeant{i}_Age'
            ])
        # Remarque: nous n'ajoutons pas de colonnes de type pour les dirigeants
        
        for col in result_columns:
            df[col] = None
            
        # Créer un fichier CSV pour les résultats
        now = datetime.now()
        date_time_str = now.strftime("%Y%m%d_%H%M%S")
        csv_output_file = f"results_{date_time_str}.csv"
        csv_path = os.path.join(os.path.dirname(output_file), csv_output_file)
        
        # Supprimer les colonnes de type de dirigeant si elles existent
        type_columns = [col for col in df.columns if col.endswith('_Type')]
        if type_columns:
            df = df.drop(columns=type_columns)
            
        # Traiter chaque ligne
        total_rows = len(df)
        matches = 0
        address_matches = 0
        
        for idx, row in df.iterrows():
            company_name = row.get('Company', '')
            address = row.get('Address', '')
            
            if pd.isna(company_name) or not company_name.strip():
                logging.warning(f"Ligne {idx+1}: Nom d'entreprise vide, ignoré")
                continue
                
            logging.info(f"Traitement de l'entreprise {idx+1}/{total_rows}: {company_name}")
            
            # Rechercher l'entreprise
            result = search_company(company_name, address, token)
            
            if result:
                matches += 1
                # Remplir les colonnes de résultats
                df.at[idx, 'SIREN'] = result['siren']
                df.at[idx, 'Nom_Raison_Sociale'] = result['nom_raison_sociale']
                df.at[idx, 'Adresse'] = result['adresse']
                df.at[idx, 'Etat_Administratif'] = result['etat_administratif']
                df.at[idx, 'Tranche_Effectif'] = result['tranche_effectif']
                df.at[idx, 'Date_Creation'] = result['date_creation']
                df.at[idx, 'Annee_Creation'] = result['annee_creation']
                df.at[idx, 'Nom_Dirigeant'] = result['nom_dirigeant']
                df.at[idx, 'Prenom_Dirigeant'] = result['prenom_dirigeant']
                df.at[idx, 'Qualite_Dirigeant'] = result['qualite_dirigeant']
                df.at[idx, 'Age_Dirigeant'] = result['age_dirigeant']
                df.at[idx, 'Match_Adresse'] = result['match_adresse']
                
                # Remplir les colonnes des dirigeants alternatifs
                autres_dirigeants = result.get('autres_dirigeants', [])
                for i, dirigeant in enumerate(autres_dirigeants, start=1):
                    if i > 5:  # Limiter à 5 dirigeants alternatifs
                        break
                    
                    df.at[idx, f'Dirigeant{i}_Nom'] = dirigeant['nom']
                    df.at[idx, f'Dirigeant{i}_Prenom'] = dirigeant['prenoms']
                    df.at[idx, f'Dirigeant{i}_Qualite'] = dirigeant['qualite']
                    df.at[idx, f'Dirigeant{i}_Age'] = dirigeant['age']
                    # Ne pas ajouter de colonne de type
                
                if result['match_adresse'] == "Oui":
                    address_matches += 1
                    logging.info(f"✓ Correspondance trouvée: SIREN {result['siren']} | Adresse: correspondante")
                else:
                    logging.info(f"! Correspondance trouvée: SIREN {result['siren']} | Adresse: différente")
            else:
                logging.warning(f"✗ Aucune correspondance trouvée pour {company_name}")
        
        # Afficher les statistiques
        if total_rows > 0:
            match_percent = (matches/total_rows*100)
            address_match_percent = (address_matches/matches*100) if matches > 0 else 0
            
            logging.info(f"\nRésumé du traitement:")
            logging.info(f"- Total des entreprises: {total_rows}")
            logging.info(f"- Entreprises trouvées: {matches}/{total_rows} ({match_percent:.1f}%)")
            logging.info(f"- Adresses correspondantes: {address_matches}/{matches} ({address_match_percent:.1f}%)")
            
        # Sauvegarder les résultats avec formatage conditionnel si openpyxl est disponible
        try:
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                df.to_excel(writer, index=False)
                
                # Vérifier si openpyxl est disponible
                if 'openpyxl' in sys.modules:
                    # Appliquer le formatage conditionnel
                    workbook = writer.book
                    worksheet = writer.sheets['Sheet1']
                    
                    # Trouver l'index de la colonne Match_Adresse
                    match_col_idx = result_columns.index('Match_Adresse') + len(required_columns) + 1
                    
                    # Appliquer une couleur verte pour les correspondances et rouge pour les non-correspondances
                    for row_idx in range(2, len(df) + 2):  # +2 car Excel commence à 1 et il y a l'en-tête
                        cell = worksheet.cell(row=row_idx, column=match_col_idx)
                        if cell.value == "Oui":
                            # Vert clair
                            cell.fill = openpyxl.styles.PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        elif cell.value == "Non":
                            # Rouge clair pour la cellule de correspondance
                            cell.fill = openpyxl.styles.PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                            
                            # Surligner toute la ligne en rouge clair si pas de correspondance d'adresse
                            light_red = openpyxl.styles.PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
                            for col_idx in range(1, len(df.columns) + 1):
                                if col_idx != match_col_idx:  # Ne pas changer la couleur de la cellule Match_Adresse
                                    worksheet.cell(row=row_idx, column=col_idx).fill = light_red
                else:
                    logging.warning("Module openpyxl non disponible, formatage conditionnel désactivé")
        except Exception as e:
            logging.warning(f"Erreur lors de l'application du formatage conditionnel: {str(e)}")
            # Sauvegarde simple sans formatage
            df.to_excel(output_file, index=False)
        
        # Sauvegarder aussi en CSV
        df.to_csv(csv_path, index=False, encoding='utf-8-sig')
        logging.info(f"Résultats sauvegardés dans {output_file} et {csv_output_file}")
        
        return True
        
    except Exception as e:
        logging.error(f"Erreur lors du traitement: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

def main():
    import argparse
    
    # Analyser les arguments
    parser = argparse.ArgumentParser(description="Enrichissement de données d'entreprise via l'API Recherche d'Entreprises")
    parser.add_argument('--input', '-i', type=str, default="data.xlsx", 
                      help="Chemin vers le fichier Excel d'entrée")
    parser.add_argument('--output', '-o', type=str, default="data_enrichi.xlsx",
                      help="Chemin vers le fichier Excel de sortie")
    parser.add_argument('--token', '-t', type=str, 
                      help="Token d'API Recherche d'Entreprises")
    parser.add_argument('--debug', '-d', action='store_true',
                      help="Active le mode debug")
    
    args = parser.parse_args()
    
    # Configurer le niveau de log
    if args.debug:
        logging.getLogger().setLevel(logging.DEBUG)
        logging.info("Mode debug activé")
    
    # Définir le token (utilisez celui fourni en argument ou celui par défaut)
    token = args.token if args.token else "14e3d8bb-190f-4fcc-a3d8-bb190fffcc2c"
    
    # Définir les chemins de fichiers
    input_path = os.path.abspath(args.input)
    output_path = os.path.abspath(args.output)
    
    logging.info(f"Traitement du fichier: {input_path}")
    logging.info(f"Le résultat sera sauvegardé dans: {output_path}")
    
    # Vérifier que les dépendances sont installées
    if 'openpyxl' not in sys.modules:
        logging.warning("Module openpyxl non trouvé. Le formatage conditionnel ne sera pas appliqué.")
        
    # Exécuter l'enrichissement
    success = enrich_excel_file(input_path, output_path, token)
    
    if success:
        logging.info("Traitement terminé avec succès")
    else:
        logging.error("Le traitement a échoué")
        sys.exit(1)

if __name__ == "__main__":
    main()